from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

DARK_BLUE = "1F4E79"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6E4F0"
LIGHT_GRAY = "F2F2F2"
GREEN_FILL = "C6EFCE"
YELLOW_FILL = "FFEB9C"
RED_FILL = "FFC7CE"
GREEN_FONT = "006100"
YELLOW_FONT = "9C6500"
RED_FONT = "9C0006"

header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=DARK_BLUE)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
title_font = Font(name="Arial", bold=True, size=16, color=DARK_BLUE)
subtitle_font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)
body_font = Font(name="Arial", size=10)
bold_font = Font(name="Arial", bold=True, size=10)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

def style_data_row(ws, row, cols, alt=False):
    fill = PatternFill("solid", fgColor=LIGHT_BLUE) if alt else PatternFill("solid", fgColor=LIGHT_GRAY)
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

def style_status(cell, status):
    if status == "Done":
        cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
        cell.font = Font(name="Arial", size=10, color=GREEN_FONT, bold=True)
    elif status == "In Progress":
        cell.fill = PatternFill("solid", fgColor=YELLOW_FILL)
        cell.font = Font(name="Arial", size=10, color=YELLOW_FONT, bold=True)
    else:
        cell.fill = PatternFill("solid", fgColor=RED_FILL)
        cell.font = Font(name="Arial", size=10, color=RED_FONT, bold=True)

def add_task_sheet(wb, name, tasks):
    ws = wb.create_sheet(name)
    headers = ["Day", "Date", "Task", "Description", "Status", "Notes", "Blockers"]
    col_widths = [6, 14, 30, 55, 14, 30, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))
    ws.row_dimensions[1].height = 25
    for r, task in enumerate(tasks, 2):
        for c, val in enumerate(task, 1):
            ws.cell(row=r, column=c, value=val)
        style_data_row(ws, r, len(headers), alt=(r % 2 == 0))
        ws.row_dimensions[r].height = 35
        style_status(ws.cell(row=r, column=5), task[4])
    return ws

# === Sheet 1: Overview ===
ws = wb.active
ws.title = "Overview"
ws.sheet_properties.tabColor = DARK_BLUE

ws.merge_cells("A1:G1")
c1 = ws.cell(row=1, column=1, value="DynLang-SLAM Progress Tracker")
c1.font = title_font
c1.alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 35

info = [
    ("Project:", "DynLang-SLAM: Dynamic-Aware Open-Vocabulary Language-Embedded 3DGS SLAM"),
    ("Student:", "Ankur"),
    ("Course:", "EEE 515: Computer Vision - Spring 2026, Arizona State University"),
    ("Timeline:", "4 Weeks (March 20 - April 17, 2026)"),
    ("GPU:", "NVIDIA RTX 5070 Laptop (12GB VRAM, Blackwell)"),
    ("Fallback:", "If dynamic masking fails -> static language GS-SLAM (Week 1-2 deliverables)"),
]
for i, (k, v) in enumerate(info, 3):
    ws.cell(row=i, column=1, value=k).font = bold_font
    ws.merge_cells(f"B{i}:G{i}")
    ws.cell(row=i, column=2, value=v).font = body_font

ws.column_dimensions["A"].width = 12
for col in "BCDEFG":
    ws.column_dimensions[col].width = 18

r = 11
ws.merge_cells(f"A{r}:G{r}")
ws.cell(row=r, column=1, value="Weekly Milestones").font = subtitle_font
ws.row_dimensions[r].height = 28

r = 13
overview_headers = ["Week", "Dates", "Phase", "Key Deliverables", "Status", "Risk", "Notes"]
ow = [8, 16, 16, 50, 14, 10, 25]
for i, w in enumerate(ow, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(overview_headers, 1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, len(overview_headers))

milestones = [
    ["Week 1", "Mar 20-24", "SLAM Backbone", "Gaussian map, differentiable rendering (gsplat), tracking, mapping, test on Replica (ATE < 2cm)", "In Progress", "Medium", "Day 1 setup complete"],
    ["Week 2", "Mar 25-29", "Language Pipeline", "CLIP extraction, SAM masks, autoencoder (768D->16D), per-Gaussian language features, open-vocab query", "Not Started", "Medium", ""],
    ["Week 3", "Mar 31-Apr 4", "Dynamic Masking", "YOLOv8-Seg, temporal filter, mask dilation, SLAM+language integration, ScanNet testing", "Not Started", "High", "Novel contribution"],
    ["Week 4", "Apr 7-11", "Eval & Paper", "Ablations, baselines, CVPR paper, presentation, submission", "Not Started", "Low", ""],
]
for i, row_data in enumerate(milestones):
    row = r + 1 + i
    for c, val in enumerate(row_data, 1):
        ws.cell(row=row, column=c, value=val)
    style_data_row(ws, row, len(overview_headers), alt=(i % 2 == 0))
    ws.row_dimensions[row].height = 40
    style_status(ws.cell(row=row, column=5), row_data[4])

# === Sheet 2: Week 1 - SLAM ===
w1_tasks = [
    [1, "2026-03-20", "Environment Setup", "Install PyTorch CUDA, gsplat, dependencies, download Replica dataset (8 scenes)", "Done", "All 11/12 checks pass, gsplat replaces diff-gaussian-rasterization", ""],
    [2, "2026-03-21", "Gaussian Map Data Structure", "GaussianMap class: positions, rotations, scales, opacities, SH coeffs, 16D lang features", "In Progress", "", ""],
    [2, "2026-03-21", "Differentiable Rendering", "Wrapper around gsplat for RGB, depth, silhouette, and language feature rendering", "Not Started", "", ""],
    [2, "2026-03-21", "Tracking Module", "6-DOF camera pose optimization via differentiable rendering against current map", "Not Started", "", ""],
    [2, "2026-03-21", "Mapping Module", "Gaussian parameter optimization + silhouette-based expansion for new regions", "Not Started", "", ""],
    [2, "2026-03-21", "SLAM Loop", "Alternating tracking-mapping pipeline with keyframe management", "Not Started", "", ""],
    [2, "2026-03-21", "Test on Replica room0", "Run SLAM on room0, target ATE < 2cm RMSE", "Not Started", "Key validation milestone", ""],
    [3, "2026-03-22", "Debug & Optimize SLAM", "Fix tracking/mapping issues, optimize for RTX 5070, profile VRAM usage", "Not Started", "", ""],
    [3, "2026-03-22", "Test More Scenes", "Run on room1, room2, office0 - verify generalization", "Not Started", "", ""],
    [4, "2026-03-23", "Keyframe Selection", "Implement overlap-based keyframe selection strategy", "Not Started", "", ""],
    [4, "2026-03-23", "Loop Closure Basics", "Simple pose graph optimization for drift correction", "Not Started", "", ""],
    [5, "2026-03-24", "Integration Testing", "Full Week 1 pipeline test on all 8 Replica scenes", "Not Started", "", ""],
    [5, "2026-03-24", "Benchmark Results", "Record ATE, PSNR, SSIM for all scenes; compare vs SplaTAM paper numbers", "Not Started", "", ""],
]
add_task_sheet(wb, "Week 1 - SLAM", w1_tasks)

# === Sheet 3: Week 2 - Language ===
w2_tasks = [
    [6, "2026-03-25", "CLIP Feature Extraction", "CLIP ViT-L/14 feature pipeline: extract 768D features per frame (every 5th frame)", "Not Started", "", ""],
    [7, "2026-03-26", "SAM Multi-Scale Masks", "Generate 3-scale masks using SAM ViT-L for hierarchical semantics", "Not Started", "SAM checkpoint already downloaded", ""],
    [8, "2026-03-27", "Autoencoder Training", "Train scene-wise autoencoder: 768D CLIP -> 16D compact features (following LangSplat)", "Not Started", "", ""],
    [9, "2026-03-28", "Language Feature Distillation", "Per-Gaussian 16D language features via L_lang cosine distance loss", "Not Started", "", ""],
    [9, "2026-03-28", "Joint Optimization", "L_total = L_rgb + L_depth + L_ssim + L_lang joint language-geometry optimization", "Not Started", "", ""],
    [10, "2026-03-29", "Open-Vocab Querying", "CLIP text encoder -> cosine similarity -> 3D localization heatmaps", "Not Started", "", ""],
    [10, "2026-03-29", "Integration Test", "Full static language GS-SLAM on Replica room0, verify mIoU", "Not Started", "FALLBACK PROJECT COMPLETE at this point", ""],
]
add_task_sheet(wb, "Week 2 - Language", w2_tasks)

# === Sheet 4: Week 3 - Dynamic ===
w3_tasks = [
    [11, "2026-03-31", "YOLOv8-Seg Detection", "Dynamic object detection: person, animal, vehicle classes", "Not Started", "", ""],
    [12, "2026-04-01", "Temporal Filter + Dilation", "3-frame sliding window consistency filter, mask dilation for boundary artifacts", "Not Started", "", ""],
    [13, "2026-04-02", "SLAM Integration", "Apply masks before tracking/mapping - exclude dynamic regions from optimization", "Not Started", "Critical integration point", ""],
    [14, "2026-04-03", "Dynamic Scene Testing", "Test on ScanNet sequences with natural human motion", "Not Started", "May need synthetic dynamic scenes from Replica", ""],
    [15, "2026-04-04", "Full Pipeline Test", "End-to-end: dynamic masking + SLAM + language features on dynamic scenes", "Not Started", "Novel contribution validated here", ""],
]
add_task_sheet(wb, "Week 3 - Dynamic", w3_tasks)

# === Sheet 5: Week 4 - Eval & Paper ===
w4_tasks = [
    [16, "2026-04-07", "Evaluation Scripts", "Implement ATE RMSE, PSNR, SSIM, mIoU metrics; dynamic robustness metric", "Not Started", "", ""],
    [17, "2026-04-08", "Baselines & Ablations", "Run SplaTAM (no semantics), OVO comparison; ablations: masking on/off, 3D vs 16D, joint vs post-hoc", "Not Started", "", ""],
    [18, "2026-04-09", "Results & Figures", "Generate comparison tables, qualitative visualizations, ablation plots", "Not Started", "", ""],
    [19, "2026-04-10", "Paper Writing", "CVPR-format: intro, related work, method, experiments, conclusion", "Not Started", "LaTeX template already set up", ""],
    [20, "2026-04-11", "Final Polish", "Paper revision, presentation slides, final submission", "Not Started", "", ""],
]
add_task_sheet(wb, "Week 4 - Eval & Paper", w4_tasks)

# === Sheet 6: Metrics ===
ws_m = wb.create_sheet("Metrics")
ws_m.sheet_properties.tabColor = "4472C4"
m_headers = ["Scene", "ATE RMSE (cm)", "PSNR (dB)", "SSIM", "mIoU (%)", "Dynamic Robust.", "FPS", "VRAM (GB)", "Notes"]
m_widths = [14, 16, 14, 10, 12, 16, 8, 12, 30]
for i, w in enumerate(m_widths, 1):
    ws_m.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(m_headers, 1):
    ws_m.cell(row=1, column=i, value=h)
style_header(ws_m, 1, len(m_headers))

scenes = ["room0", "room1", "room2", "office0", "office1", "office2", "office3", "office4"]
for i, scene in enumerate(scenes):
    r = i + 2
    ws_m.cell(row=r, column=1, value=scene)
    style_data_row(ws_m, r, len(m_headers), alt=(i % 2 == 0))
    ws_m.row_dimensions[r].height = 22

r_avg = len(scenes) + 2
ws_m.cell(row=r_avg, column=1, value="AVERAGE").font = bold_font
ws_m.cell(row=r_avg, column=1).border = thin_border
for c in range(2, 9):
    col_letter = get_column_letter(c)
    formula = f"=IF(COUNT({col_letter}2:{col_letter}9)>0,AVERAGE({col_letter}2:{col_letter}9),\"\")"
    cell = ws_m.cell(row=r_avg, column=c, value=formula)
    cell.font = bold_font
    cell.border = thin_border

# Tab colors
for name, color in [("Week 1 - SLAM", "2E75B6"), ("Week 2 - Language", "548235"), ("Week 3 - Dynamic", "BF8F00"), ("Week 4 - Eval & Paper", "C00000")]:
    wb[name].sheet_properties.tabColor = color

out = r"C:\Users\ankur\Desktop\research\DynLang-SLAM\DynLang_SLAM_Progress_Tracker.xlsx"
wb.save(out)
print(f"Saved to {out}")
